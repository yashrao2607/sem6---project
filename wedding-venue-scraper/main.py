"""
Indian Wedding Destination Intelligence Pipeline — main entry point.

Usage:
    python main.py [--sources wedmegood venuelook shaadisaga]
                   [--states "Rajasthan" "Goa"]
                   [--skip-scraping]
                   [--skip-analytics]
                   [--invalidate-cache]
                   [--input-csv path/to/existing.csv]

Phases:
    1. Scraping     — fetch venues from Tier 1 sources per city
    2. Cleaning     — normalize, validate, deduplicate
    3. Export       — CSV + Excel with pivot sheets
    4. Analytics    — 10 static charts + interactive dashboard
    5. Report       — markdown summary with advanced insights
"""
import argparse
import json
import logging
import sys
from pathlib import Path

import pandas as pd

# ── Project root on sys.path ──────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

from config.settings import (
    FINAL_CSV, FINAL_XLSX,
    SOURCE_WEDMEGOOD, SOURCE_VENUELOOK, SOURCE_SHAADISAGA,
    RAW_DIR, PROCESSED_DIR, DATASET_DIR,
)
from src.scrapers.wedmegood   import WedMeGoodScraper
from src.scrapers.venuelook   import VenueLookScraper
from src.scrapers.shaadisaga  import ShaadiSagaScraper
from src.cleaning.normalizer  import normalize_record
from src.cleaning.validator   import validate_records, apply_median_imputation
from src.cleaning.deduplicator import deduplicate, report_duplicate_rate
from src.analytics.charts    import generate_all_charts
from src.analytics.dashboard import build_dashboard
from src.analytics.report    import generate_report

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(name)s - %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(ROOT / "pipeline.log", encoding="utf-8"),
    ]
)
# Ensure stdout can handle unicode on Windows
if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
logger = logging.getLogger("pipeline")

# ── Source registry ───────────────────────────────────────────────────────────
SCRAPER_MAP = {
    SOURCE_WEDMEGOOD:  WedMeGoodScraper,
    SOURCE_VENUELOOK:  VenueLookScraper,
    SOURCE_SHAADISAGA: ShaadiSagaScraper,
}


# ── Phase 1: Scraping ─────────────────────────────────────────────────────────

def phase_scrape(sources: list[str], target_states: list[str]) -> list[dict]:
    cities_config = json.loads((ROOT / "config" / "cities.json").read_text())
    all_raw: list[dict] = []

    for source_name in sources:
        scraper_cls = SCRAPER_MAP.get(source_name)
        if not scraper_cls:
            logger.warning("Unknown source '%s' — skipping", source_name)
            continue

        scraper = scraper_cls()
        logger.info("=== Scraping source: %s ===", source_name)

        # Slug key for this source (wmg_slug / vl_slug / fallback to city key)
        slug_field = {
            SOURCE_WEDMEGOOD:  "wmg_slug",
            SOURCE_VENUELOOK:  "vl_slug",
            SOURCE_SHAADISAGA: "wmg_slug",  # ShaadiSaga uses same pattern as WedMeGood
        }.get(source_name, "wmg_slug")

        for state, state_data in cities_config["states"].items():
            if target_states and state not in target_states:
                continue
            for city_key, city_data in state_data["cities"].items():
                city_display = city_data["display"]
                city_slug = city_data.get(slug_field) or city_data.get("wmg_slug") or city_key
                if not city_slug:
                    logger.info("  -- %s / %s not available on %s — skipping",
                                state, city_display, source_name)
                    continue
                try:
                    venues = scraper.scrape_city(city_slug, state, city_display)
                    all_raw.extend(venues)
                    logger.info("  OK %s / %s -> %d venues", state, city_display, len(venues))
                except Exception as exc:
                    logger.error("  FAIL %s / %s: %s", state, city_display, exc)

        # Save per-source raw dump
        RAW_DIR.mkdir(parents=True, exist_ok=True)
        raw_df = pd.DataFrame(all_raw)
        raw_path = RAW_DIR / f"{source_name}_raw.csv"
        raw_df.to_csv(raw_path, index=False, encoding="utf-8-sig")
        logger.info("Raw dump saved → %s (%d records)", raw_path, len(raw_df))

    return all_raw


# ── Phase 2: Cleaning ─────────────────────────────────────────────────────────

def phase_clean(raw_records: list[dict], cities_config: dict) -> pd.DataFrame:
    # Build area lookup: {city_slug: {area_lower: area_display}}
    area_lookups: dict = {}
    for state_data in cities_config["states"].values():
        for city_slug, city_data in state_data["cities"].items():
            city_display = city_data["display"]
            area_lookups[city_display.lower()] = {
                area.lower(): area for area in city_data.get("areas", [])
            }

    # 1. Normalize
    logger.info("Normalizing %d raw records…", len(raw_records))
    normalized: list[dict] = []
    for raw in raw_records:
        city_key = (raw.get("city") or "").lower()
        lookup = area_lookups.get(city_key, {})
        try:
            normalized.append(normalize_record(raw, lookup))
        except Exception as exc:
            logger.debug("Normalization failed for '%s': %s", raw.get("venue_name"), exc)

    # 2. Within-source dedup before cross-source
    from src.cleaning.deduplicator import deduplicate_within_source
    normalized = deduplicate_within_source(normalized)
    logger.info("After within-source dedup: %d records", len(normalized))

    # 3. Validate
    valid, flagged, report = validate_records(normalized)
    logger.info("Validation: %d valid, %d flagged, %d dropped", len(valid), len(flagged), report.dropped)
    print(report)

    # 4. Cross-source dedup
    all_records = valid + flagged
    deduped, dup_count = deduplicate(all_records)
    dup_rate = report_duplicate_rate(len(all_records), len(deduped))
    logger.info("Cross-source dedup: removed %d duplicates (%.1f%%)", dup_count, dup_rate * 100)

    # 5. Median imputation for missing prices
    deduped = apply_median_imputation(deduped)

    df = pd.DataFrame(deduped)
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    processed_path = PROCESSED_DIR / "venues_clean.csv"
    df.to_csv(processed_path, index=False, encoding="utf-8-sig")
    logger.info("Cleaned dataset saved → %s (%d records)", processed_path, len(df))
    return df


# ── Phase 3: Export ───────────────────────────────────────────────────────────

def phase_export(df: pd.DataFrame):
    DATASET_DIR.mkdir(parents=True, exist_ok=True)

    # CSV
    df.to_csv(FINAL_CSV, index=False, encoding="utf-8-sig")
    logger.info("Final CSV → %s", FINAL_CSV)

    # Excel with multiple sheets
    _export_excel(df)


def _export_excel(df: pd.DataFrame):
    try:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        # Sheet 1: Raw data
        ws_data = wb.active
        ws_data.title = "Venue Data"
        _write_df_to_ws(ws_data, df)

        # Sheet 2: State pivot
        state_pivot = df.groupby("state").agg(
            venue_count=("venue_id", "count"),
            avg_min_price=("min_price", "mean"),
            median_min_price=("min_price", "median"),
            avg_rating=("rating", "mean"),
        ).reset_index()
        ws_state = wb.create_sheet("State Summary")
        _write_df_to_ws(ws_state, state_pivot)

        # Sheet 3: City pivot
        city_pivot = df.groupby(["state", "city"]).agg(
            venue_count=("venue_id", "count"),
            avg_min_price=("min_price", "mean"),
            avg_rating=("rating", "mean"),
        ).reset_index()
        ws_city = wb.create_sheet("City Summary")
        _write_df_to_ws(ws_city, city_pivot)

        # Sheet 4: Data quality summary
        quality_rows = [
            ["Field", "Non-Null Count", "Fill Rate"],
        ]
        for col in df.columns:
            non_null = df[col].notna().sum()
            rate = non_null / len(df) if len(df) else 0
            quality_rows.append([col, non_null, f"{rate:.1%}"])
        ws_quality = wb.create_sheet("Data Quality")
        for row in quality_rows:
            ws_quality.append(row)

        wb.save(str(FINAL_XLSX))
        logger.info("Final Excel → %s", FINAL_XLSX)
    except ImportError:
        logger.warning("openpyxl not installed — skipping Excel export. "
                       "Install with: pip install openpyxl")


def _write_df_to_ws(ws, df: pd.DataFrame):
    from openpyxl.styles import Font, PatternFill
    # Header row
    ws.append(list(df.columns))
    header_row = ws[1]
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    # Data rows
    for row in df.itertuples(index=False):
        ws.append(list(row))
    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)


# ── Phase 4 & 5: Analytics ────────────────────────────────────────────────────

def phase_analytics(df: pd.DataFrame):
    logger.info("=== Generating charts ===")
    chart_paths = generate_all_charts(df)
    generated = [p for p in chart_paths if p]
    logger.info("Generated %d / %d charts", len(generated), len(chart_paths))

    logger.info("=== Building interactive dashboard ===")
    dash_path = build_dashboard(df)
    if dash_path:
        logger.info("Dashboard → %s", dash_path)

    logger.info("=== Generating summary report ===")
    report_path = generate_report(df)
    logger.info("Report → %s", report_path)


# ── CLI ───────────────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="Indian Wedding Venue Intelligence Pipeline"
    )
    parser.add_argument(
        "--sources", nargs="+",
        default=[SOURCE_WEDMEGOOD, SOURCE_VENUELOOK],
        choices=list(SCRAPER_MAP.keys()),
        help="Sources to scrape (default: wedmegood venuelook)"
    )
    parser.add_argument(
        "--states", nargs="*", default=[],
        help="Restrict to specific states (default: all)"
    )
    parser.add_argument(
        "--skip-scraping", action="store_true",
        help="Skip scraping; load from data/processed/venues_clean.csv"
    )
    parser.add_argument(
        "--skip-analytics", action="store_true",
        help="Skip chart and dashboard generation"
    )
    parser.add_argument(
        "--invalidate-cache", action="store_true",
        help="Invalidate the URL cache before running"
    )
    parser.add_argument(
        "--input-csv", type=Path, default=None,
        help="Path to an existing cleaned CSV to use instead of scraping"
    )
    return parser.parse_args()


def main():
    args = parse_args()

    if args.invalidate_cache:
        from src.scrapers.base import CacheDB
        logger.info("Invalidating URL cache…")
        CacheDB().invalidate()

    cities_config = json.loads((ROOT / "config" / "cities.json").read_text())

    # ── Load or scrape ────────────────────────────────────────────────────────
    if args.input_csv and args.input_csv.exists():
        logger.info("Loading dataset from %s", args.input_csv)
        df = pd.read_csv(args.input_csv)
    elif args.skip_scraping:
        clean_path = PROCESSED_DIR / "venues_clean.csv"
        if not clean_path.exists():
            logger.error("No cleaned data found at %s. Run without --skip-scraping first.", clean_path)
            sys.exit(1)
        logger.info("Loading existing cleaned dataset from %s", clean_path)
        df = pd.read_csv(clean_path)

    if args.states and not df.empty:
        logger.info("Filtering dataset for states: %s", ", ".join(args.states))
        df = df[df["state"].isin(args.states)]
        if df.empty:
            logger.error("No records found for states: %s", args.states)
            sys.exit(1)
    else:
        raw_records = phase_scrape(args.sources, args.states)
        if not raw_records:
            logger.error("No records scraped — check source availability and network access.")
            sys.exit(1)
        df = phase_clean(raw_records, cities_config)
        phase_export(df)

    logger.info("Dataset loaded: %d venues across %d states",
                len(df), df["state"].nunique() if "state" in df else 0)

    if not args.skip_analytics:
        phase_analytics(df)

    logger.info("Pipeline complete.")
    return df


if __name__ == "__main__":
    main()
